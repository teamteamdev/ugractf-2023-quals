import json
import os.path
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side

from kyzylborda_lib.generator import get_attachments_dir
from kyzylborda_lib.secrets import get_flag


def render_text(text):
    with open("seven-plus.json") as f:
        font = json.load(f)

    lines = [""] * max(font["lineHeight"], max(glyph["offset"] + len(glyph["pixels"]) for glyph in font["glyphs"].values()))

    for c in text:
        glyph = font["glyphs"][c]
        for i in range(len(lines)):
            j = i - glyph["offset"]
            if 0 <= j < len(glyph["pixels"]):
                line = "".join("#" if c else " " for c in glyph["pixels"][j])
            else:
                line = " " * len(glyph["pixels"][0])
            lines[i] += line + " "

    while lines[-1].strip() == "":
        lines.pop()

    return lines


def generate():
    lines = render_text(get_flag())
    lines = [" " + line for line in lines]

    width = len(lines[0])
    lines.insert(0, " " * width)
    lines.append(" " * width)
    height = len(lines)

    wb = openpyxl.Workbook()
    ws = wb.active

    for i in range(1, height + 1):
        for j in range(1, width + 1):
            cell = ws.cell(row=i, column=j)
            if lines[i - 1][j - 1] == "#":
                value = "=ROW(" + cell.coordinate + ")*COLUMN(" + cell.coordinate + ")"
            else:
                value = i * j
            cell.value = value
            cell.border = Border(
                left=Side(style="thick"),
                right=Side(style="thick"),
                top=Side(style="thick"),
                bottom=Side(style="thick")
            )
            cell.font = Font(color="ffffff")

    for i in range(1, height + 1):
        ws.row_dimensions[i].height = 25
    for j in range(1, width + 1):
        ws.column_dimensions[ws.cell(row=1, column=j).coordinate[:-1]].width = 5

    wb.save(filename=os.path.join(get_attachments_dir(), "picture.xlsx"))
